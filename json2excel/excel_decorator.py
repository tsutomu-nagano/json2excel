import yaml
import openpyxl
import pandas as pd
import re
import json
import sys
import copy
import shutil
import utils

class ExcelDecorator():

    @staticmethod
    def add_border(cell, position, side):
        if position == "all":
            border_ = {"left": side ,"right": side, "top": side, "bottom": side}
        else:
            border_ = {"left": cell.border.left ,"right": cell.border.right, "top": cell.border.top, "bottom": cell.border.bottom}
            border_[position] = side
        
        cell.border = openpyxl.styles.Border(** border_)

    def get_sheet_names(wb, info):

        if "re" in info:
            name_match = lambda s : re.match(info["re"], s)

        if "text" in info:
            name_match = lambda s : info["text"] == s

        return([s for s in wb.sheetnames if name_match(s)])


    @classmethod
    def decoration(cls, config, dest):
        
        # Hyperlinkなどの装飾処理
        with open(config, 'r') as ymlf:
            deco = yaml.safe_load(ymlf)

        wb = openpyxl.load_workbook(dest)

        if "hyperlink" in deco:
            for hyperlink in deco["hyperlink"]:

                from_sheet_info = hyperlink["from"]["sheet"]
                from_address = hyperlink["from"]["address"]
                to_sheet_info = hyperlink["to"]["sheet"]
                to_address = hyperlink["to"]["address"]

                for r in utils.a1_to_range(from_address):

                    cell = wb[from_sheet_info["text"]][r["address"]]

                    if to_sheet_info["text"] == "$.":
                        to_sheet = cell.value


                    if to_sheet in wb.sheetnames:

                        cell.hyperlink = f"#{to_sheet}!{to_address}"
                        cell.font = openpyxl.styles.Font(color="0000FF", underline="single")

                        if "direction" in hyperlink:
                            direction = hyperlink["direction"]
                            if direction == "both":
                                
                                cell = wb[to_sheet][to_address]
                                cell.hyperlink = f"#{from_sheet_info['text']}!{r['address']}"
                                cell.font = openpyxl.styles.Font(color="0000FF", underline="single")

        if "border" in deco:
            for border in deco["border"]:
                
                sheet_names = ExcelDecorator.get_sheet_names(wb, border["sheet"])
                draws = border["draw"]

                for sheet_name in sheet_names:

                    ranges = utils.a1_to_range(utils.convert_auto_range(border["address"], dest, sheet_name))

                    for draw in draws:
                        draw_ = draw.split("-")
                        target = draw_[0]
                        style = draw_[1]

                        side_ = openpyxl.styles.Side(border_style = style, color='000000')

                        if target == "grid":
                            [ExcelDecorator.add_border(wb[sheet_name][r["address"]], "all", side_) for r in ranges]

                        if target == "around":
                            arounds = utils.around_ranges(ranges)

                            [ExcelDecorator.add_border(wb[sheet_name][r["address"]], position, side_) for position, ranges in arounds.items() for r in ranges]

                        if target in ["bottom", "top", "left", "right"]:
                            [ExcelDecorator.add_border(wb[sheet_name][r["address"]], target, side_) for r in ranges]

        if "fill" in deco:

            for fill in deco["fill"]:
                sheet_names = ExcelDecorator.get_sheet_names(wb, fill["sheet"])

                for sheet_name in sheet_names:

                    ranges = utils.a1_to_range(utils.convert_auto_range(fill["address"], dest, sheet_name))

                    for color in fill["color"]:
                        fill_ = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=color["value"])

                        text_match = lambda t : True
                        if "text" in color:
                            if "match" in color["text"]:
                                text_match = lambda t : t in color["text"]["match"]

                        for r in [r for r in ranges if text_match(wb[sheet_name][r["address"]].value)]:
                            wb[sheet_name][r["address"]].fill = fill_


        wb.save(dest)
        wb.close()



if __name__ == "__main__":
    args = sys.argv

    src = args[1]
    config = args[2]

    ExcelDecorator.decoration(config, src)



